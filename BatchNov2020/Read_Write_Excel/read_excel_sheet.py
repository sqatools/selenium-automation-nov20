import openpyxl
"""
Install module using pip
pip install openpyxl
"""
"""
filepath = 'test_data.xlsx'

wb = openpyxl.load_workbook(filepath)
asheet = wb.active
sheet = wb.sheetnames
print(sheet)

print(asheet.cell(1, 1).value)


for i in range(1, 10):
    print(asheet.cell(i, 1).value)

"""


def read_excel_data(filepath, row, col):
    wb = openpyxl.load_workbook(filepath)
    a_sheet = wb.active
    return a_sheet.cell(row, col).value


#username = read_excel_data('test_data.xlsx', 14, 1)
#password = read_excel_data('test_data.xlsx', 14, 2)

#(username, password)


def read_excel_data_by_cellname(filepath, cellname):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    print(sheet[cellname].value)
    print(sheet["A1"].value)


#read_excel_data_by_cellname('test_data.xlsx', "I8")


def read_complete_col(filepath, cell):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    allcell = sheet[cell]
    for block in allcell:
        print(block.value)


#read_complete_col("test_data.xlsx", "A")
#read_complete_col("test_data.xlsx", "B")

def read_employee_details(empid, filepath="Employee_details.xlsx"):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    print(f"Name : {sheet['A'+str(empid)].value},"
          f"Email : {sheet['B'+str(empid)].value}, "
          f"Salary : {sheet['C'+str(empid)].value}, "
          f"Age : {sheet['D'+str(empid)].value}")

#read_employee_details(8)

def remove_employee_data(empid, filepath="Employee_details.xlsx"):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    adata = sheet['A']
    new_data = []
    for i in range(0, len(adata)):
        print(i)
        print(adata[i].value)
        if i == empid-1:
            continue
        else:
            #print(i.value)
            new_data.append(adata[i].value)

    print(new_data)

     # Update existing excel sheet
    wb2 = openpyxl.Workbook()
    usheet= wb2.active
    count = 1
    for data in new_data:
        usheet['A'+str(count)] = data
        count += 1
    wb2.save(filepath)

remove_employee_data(13)